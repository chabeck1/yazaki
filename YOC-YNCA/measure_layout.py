#!/usr/bin/env python3
"""
Measures column widths and row heights for key columns in a source SOV file,
including the blank "form" region (first 14 rows) of an Excel sheet.

This script finds the sizes (in Excel units) of the first 14 rows—so you can
replicate those blank cells for manual entry—and also locates and measures
the main data table header for reference.

Usage:
  python measure_layout.py <path_to_sov_file.pdf_or_xlsx>

Dependencies:
  pip install pdfplumber openpyxl
"""
import argparse
import os
import sys
from collections import Counter, OrderedDict

# --- Configuration ---
# Define the headers we want to find and measure the columns for.
TARGET_HEADERS = {
    'SOV Level':    'sov level',
    'Description':  'description',
    'Part Type':    'part type',
    'Part Name':    'part name',
    'Part Number':  'part number',
    'Qty':          'qty',
    'Drawing No':   'draw',
    'Rev.':         'rev',
}

# --- Measurement Logic ---

def measure_pdf(path: str) -> dict:
    import pdfplumber
    results = {}
    print(f"📄 Analyzing PDF: {path}")

    with pdfplumber.open(path) as pdf:
        page = pdf.pages[0]
        tables = page.find_tables(table_settings={
            'vertical_strategy': 'lines_strict',
            'horizontal_strategy': 'lines_strict',
        })
        if not tables:
            raise RuntimeError("No table found on the first page of the PDF.")
        table = tables[0]

        # measure header row
        header_row = table.rows[0]
        header_map = {}
        for cell in header_row.cells:
            if not cell or not cell.get('text'): continue
            x0, _, x1, _ = cell['bbox']
            text = cell['text'].strip().lower()
            for name, keyword in TARGET_HEADERS.items():
                if keyword in text:
                    header_map[name] = {'width': x1 - x0}

        # most common row height
        row_heights = [round(r.height, 1) for r in table.rows if r.height]
        common_h = Counter(row_heights).most_common(1)[0][0]
        for name in header_map:
            header_map[name]['height'] = common_h
    return header_map


def measure_blank_region(sheet, num_rows: int = 14) -> dict:
    """
    Measure the widths of all columns and heights of the first `num_rows` rows
    in the Excel sheet, so you can replicate the blank "form" region.

    Returns:
        { 'columns': OrderedDict{col_letter: width},
          'rows':    OrderedDict{row_index: height} }
    """
    import openpyxl
    results = {'columns': OrderedDict(), 'rows': OrderedDict()}
    default_width  = sheet.sheet_format.defaultColWidth or 8.43
    default_height = sheet.sheet_format.defaultRowHeight or 15.0

    # measure columns across the entire sheet
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cd = sheet.column_dimensions.get(col_letter)
        w = cd.width if cd and cd.width is not None else default_width
        results['columns'][col_letter] = w

    # measure the first `num_rows` row heights
    for r in range(1, num_rows + 1):
        rd = sheet.row_dimensions.get(r)
        h = rd.height if rd and rd.height is not None else default_height
        results['rows'][r] = h

    return results


def print_blank_region(msr: dict):
    print("\n--- Blank Region Measurements (Rows 1–14) ---")
    print("Columns (width in characters):")
    for col, w in msr['columns'].items():
        print(f"  {col}: {w:.2f}")
    print("\nRows (height in points):")
    for row, h in msr['rows'].items():
        print(f"  Row {row}: {h:.2f}")


def measure_xlsx(path: str) -> dict:
    import openpyxl
    print(f"📊 Analyzing Excel file: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active

    # first, measure blank/form region
    blank_msr = measure_blank_region(sheet)
    print_blank_region(blank_msr)

    # then find and measure the main data header
    best_score = 0
    header_row_idx = -1
    found_headers = {}
    print("\nScanning for the main data table header (for reference)...")
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        score = 0
        headers = {}
        for col_idx, val in enumerate(row, 1):
            if not val: continue
            text = str(val).strip().lower()
            for name, kw in TARGET_HEADERS.items():
                if kw in text:
                    score += 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    headers[name if name == 'SOV Level' else name] = {'col': col_letter}
        if score > best_score:
            best_score = score
            header_row_idx = row_idx
            found_headers = headers

    if best_score < 3:
        print(f"Warning: only found {best_score} matching headers; skipping detailed header measurement.")
        return blank_msr

    print(f"✅ Data header found at row {header_row_idx} with score {best_score}.")

    # now measure widths/heights for those header columns
    results = {}
    row_dim = sheet.row_dimensions.get(header_row_idx)
    header_height = row_dim.height if row_dim and row_dim.height is not None else sheet.sheet_format.defaultRowHeight
    for name, info in found_headers.items():
        col = info['col']
        col_dim = sheet.column_dimensions.get(col)
        width = col_dim.width if col_dim and col_dim.width is not None else sheet.sheet_format.defaultColWidth
        results[name] = {'width': width, 'height': header_height}

    return {'blank_region': blank_msr, 'header': results}


def print_results(sizes: dict, is_pdf: bool):
    """Prints header measurements only."""
    header = sizes.get('header', sizes)
    if not header:
        print("No header measurements available.")
        return
    print("\n--- Header Layout Measurement Results ---")
    unit = "chars" if not is_pdf else "pt"
    print(f"Units: width in {unit}, height in pt")
    print('-'*40)
    print(f"{'Component':<15} | {'Width':>6} | {'Height':>6}")
    print('-'*40)
    for name, v in header.items():
        w = v['width']; h = v['height']
        print(f"{name:<15} | {w:>6.2f} | {h:>6.2f}")
    print('-'*40)


def main():
    parser = argparse.ArgumentParser(description='Measure layout from an existing SOV file.')
    parser.add_argument('file', help='Path to the source .pdf or .xlsx SOV file')
    args = parser.parse_args()

    fp = args.file
    if not os.path.isfile(fp):
        print(f"Error: File not found at '{fp}'")
        sys.exit(1)

    ext = os.path.splitext(fp)[1].lower()
    try:
        if ext == '.pdf':
            hdr = measure_pdf(fp)
            print_results(hdr, is_pdf=True)
        elif ext in ('.xlsx', '.xlsm'):
            all_msrs = measure_xlsx(fp)
            # print only header measurements here; blank region printed inside measure_xlsx
            print_results(all_msrs, is_pdf=False)
        else:
            print(f"Error: Unsupported file type '{ext}'. Use .pdf or .xlsx.")
            sys.exit(1)
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
